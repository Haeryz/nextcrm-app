#!/usr/bin/env python3
"""Extract a cleaned customer catalog payload from the source XLSX workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import posixpath
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "PART CATALOG ALL.xlsx"
DEFAULT_OUTPUT = ROOT / "lib" / "catalog" / "generated" / "catalog-clean.json"
DEFAULT_IMAGE_DIR = ROOT / "public" / "catalog" / "images"


@dataclass(frozen=True)
class DrawingImage:
    media_path: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
    return slug.strip("-")[:80]


def hash_id(parts: list[Any]) -> str:
    raw = "|".join("" if part is None else str(part) for part in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def row_cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) and row[index] is not None else ""


def row_looks_like_header(row: tuple[Any, ...]) -> bool:
    return any(normalize_text(value).lower() == "description" for value in row)


def row_has_catalog_data(row: tuple[Any, ...]) -> bool:
    return bool(
        normalize_text(row_cell(row, 2))
        or normalize_text(row_cell(row, 1))
        or normalize_text(row_cell(row, 4))
    )


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def zip_text(workbook_zip: zipfile.ZipFile, path: str) -> str:
    try:
        return workbook_zip.read(path).decode("utf-8")
    except KeyError:
        return ""


def xml_attrs(xml: str, element_name: str) -> list[dict[str, str]]:
    attrs: list[dict[str, str]] = []
    for match in re.finditer(rf"<{element_name}\b([^>]*)/?>", xml):
        attrs.append(dict(re.findall(r'([\w:.-]+)="([^"]*)"', match.group(1))))
    return attrs


def parse_relationships(xml: str) -> dict[str, dict[str, str]]:
    relationships: dict[str, dict[str, str]] = {}
    for attrs in xml_attrs(xml, "Relationship"):
        rel_id = attrs.get("Id")
        if rel_id:
            relationships[rel_id] = attrs
    return relationships


def resolve_target(source_path: str, target: str | None) -> str:
    if not target:
        return ""
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_path), target))


def workbook_sheet_paths(workbook_zip: zipfile.ZipFile) -> dict[str, str]:
    workbook_xml = zip_text(workbook_zip, "xl/workbook.xml")
    workbook_rels = parse_relationships(
        zip_text(workbook_zip, "xl/_rels/workbook.xml.rels")
    )
    sheets: dict[str, str] = {}
    for attrs in xml_attrs(workbook_xml, "sheet"):
        rel = workbook_rels.get(attrs.get("r:id", ""))
        if attrs.get("name") and rel and rel.get("Target"):
            sheets[attrs["name"]] = resolve_target("xl/workbook.xml", rel["Target"])
    return sheets


def drawing_path_for_sheet(workbook_zip: zipfile.ZipFile, worksheet_path: str) -> str:
    rels_path = posixpath.join(
        posixpath.dirname(worksheet_path),
        "_rels",
        f"{posixpath.basename(worksheet_path)}.rels",
    )
    rels = parse_relationships(zip_text(workbook_zip, rels_path))
    drawing_rel = next(
        (rel for rel in rels.values() if "/drawing" in rel.get("Type", "")),
        None,
    )
    return resolve_target(worksheet_path, drawing_rel.get("Target") if drawing_rel else None)


def parse_drawing_images(workbook_zip: zipfile.ZipFile, drawing_path: str) -> list[DrawingImage]:
    if not drawing_path:
        return []

    drawing_xml = zip_text(workbook_zip, drawing_path)
    rels_path = posixpath.join(
        posixpath.dirname(drawing_path),
        "_rels",
        f"{posixpath.basename(drawing_path)}.rels",
    )
    rels = parse_relationships(zip_text(workbook_zip, rels_path))
    anchors = re.findall(
        r"<xdr:(?:oneCellAnchor|twoCellAnchor)\b[\s\S]*?</xdr:(?:oneCellAnchor|twoCellAnchor)>",
        drawing_xml,
    )
    images: list[DrawingImage] = []

    for anchor in anchors:
        if "<xdr:pic" not in anchor:
            continue

        row_matches = [int(row) + 1 for row in re.findall(r"<xdr:row>(\d+)</xdr:row>", anchor)]
        col_matches = [int(col) + 1 for col in re.findall(r"<xdr:col>(\d+)</xdr:col>", anchor)]
        embed_match = re.search(r'r:embed="([^"]+)"', anchor)
        if not row_matches or not embed_match:
            continue

        rel = rels.get(embed_match.group(1))
        if not rel or not rel.get("Target"):
            continue

        images.append(
            DrawingImage(
                media_path=resolve_target(drawing_path, rel["Target"]),
                start_row=row_matches[0],
                end_row=row_matches[1] if len(row_matches) > 1 else row_matches[0],
                start_col=col_matches[0] if col_matches else 0,
                end_col=col_matches[1] if len(col_matches) > 1 else (col_matches[0] if col_matches else 0),
            )
        )

    return sorted(images, key=lambda image: (image.start_row, image.start_col, image.media_path))


def item_from_row(sheet_name: str, row: tuple[Any, ...], row_index: int, image_key: str | None) -> dict[str, Any]:
    excel_row_number = row_index + 1
    illustration = normalize_text(row_cell(row, 0))
    part_number = normalize_text(row_cell(row, 1))
    description = normalize_text(row_cell(row, 2))
    quantity_value = row_cell(row, 3)
    catalog_part_number = normalize_text(row_cell(row, 4))
    price_or_remark = row_cell(row, 6)
    price = int(price_or_remark) if is_number(price_or_remark) else None
    remark = "" if is_number(price_or_remark) else normalize_text(price_or_remark)
    item_id = f"{slugify(sheet_name)}-{hash_id([
        sheet_name,
        excel_row_number,
        illustration,
        part_number,
        description,
        catalog_part_number,
    ])}"
    search_text = " ".join(
        value
        for value in [
            sheet_name,
            illustration,
            part_number,
            catalog_part_number,
            description,
            remark,
        ]
        if value
    ).lower()

    return {
        "id": item_id,
        "machine": sheet_name.strip(),
        "rowNumber": excel_row_number,
        "illustration": illustration or None,
        "partNumber": part_number or None,
        "catalogPartNumber": catalog_part_number or None,
        "description": description or catalog_part_number or part_number or "Catalogue item",
        "quantity": None if quantity_value == "" else str(quantity_value),
        "price": price,
        "remark": remark or None,
        "imageKey": image_key,
        "searchText": search_text,
    }


def assign_sheet_images(rows: list[tuple[int, tuple[Any, ...]]], images: list[DrawingImage]) -> dict[int, str]:
    if not rows or not images:
        return {}

    row_numbers = [row_number for row_number, _row in rows]
    image_by_row = {image.start_row: image.media_path for image in images}
    assignments: dict[int, str] = {}

    # Some sheets place the picture gallery below the table. In that layout, row
    # proximity is meaningless, so use visual order and repeat the last image if
    # there are fewer images than products.
    if min(image.start_row for image in images) > max(row_numbers):
        for index, row_number in enumerate(row_numbers):
            assignments[row_number] = images[min(index, len(images) - 1)].media_path
        return assignments

    sorted_images = sorted(images, key=lambda image: (image.start_row, image.start_col))
    image_index = 0
    last_image = sorted_images[0].media_path

    for row_number in row_numbers:
        direct = (
            image_by_row.get(row_number)
            or image_by_row.get(row_number - 1)
            or image_by_row.get(row_number + 1)
        )
        while image_index < len(sorted_images) and sorted_images[image_index].start_row <= row_number:
            last_image = sorted_images[image_index].media_path
            image_index += 1
        assignments[row_number] = direct or last_image

    return assignments


def image_output_name(media_path: str, bytes_value: bytes) -> str:
    digest = hashlib.sha1(media_path.encode("utf-8") + bytes_value[:256]).hexdigest()[:12]
    extension = Path(media_path).suffix.lower() or ".jpg"
    return f"{digest}{extension}"


def extract_catalog(source: Path, output: Path, image_dir: Path) -> dict[str, Any]:
    if not source.exists():
        raise FileNotFoundError(f"Workbook not found: {source}")

    output.parent.mkdir(parents=True, exist_ok=True)
    if image_dir.exists():
        shutil.rmtree(image_dir)
    image_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source, data_only=True, read_only=True)
    items: list[dict[str, Any]] = []
    used_media_paths: set[str] = set()

    with zipfile.ZipFile(source) as workbook_zip:
        sheet_paths = workbook_sheet_paths(workbook_zip)
        sheet_images = {
            sheet_name: parse_drawing_images(
                workbook_zip,
                drawing_path_for_sheet(workbook_zip, worksheet_path),
            )
            for sheet_name, worksheet_path in sheet_paths.items()
        }

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            header_index = next(
                (index for index, row in enumerate(rows) if row_looks_like_header(row)),
                -1,
            )
            data_start = header_index + 2 if header_index >= 0 else 0
            catalog_rows = [
                (row_index + 1, rows[row_index])
                for row_index in range(data_start, len(rows))
                if row_has_catalog_data(rows[row_index])
            ]
            image_assignments = assign_sheet_images(catalog_rows, sheet_images.get(sheet_name, []))

            for excel_row_number, row in catalog_rows:
                image_key = image_assignments.get(excel_row_number)
                if image_key:
                    used_media_paths.add(image_key)
                items.append(item_from_row(sheet_name, row, excel_row_number - 1, image_key))

        image_records: list[dict[str, str]] = []
        for media_path in sorted(used_media_paths):
            try:
                bytes_value = workbook_zip.read(media_path)
            except KeyError:
                continue

            filename = image_output_name(media_path, bytes_value)
            output_path = image_dir / filename
            output_path.write_bytes(bytes_value)
            mime_type = mimetypes.guess_type(filename)[0] or "image/jpeg"
            image_records.append(
                {
                    "key": media_path,
                    "file": str(output_path.relative_to(ROOT)).replace("\\", "/"),
                    "mimeType": mime_type,
                }
            )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": str(source.relative_to(ROOT)) if source.is_relative_to(ROOT) else str(source),
        "stats": {
            "items": len(items),
            "images": len(image_records),
            "itemsWithoutImage": sum(1 for item in items if not item["imageKey"]),
        },
        "images": image_records,
        "items": items,
    }
    output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--image-dir", type=Path, default=DEFAULT_IMAGE_DIR)
    args = parser.parse_args()

    payload = extract_catalog(args.source.resolve(), args.output.resolve(), args.image_dir.resolve())
    print(
        "Extracted "
        f"{payload['stats']['items']} catalogue items, "
        f"{payload['stats']['images']} images, "
        f"{payload['stats']['itemsWithoutImage']} items without image."
    )
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
