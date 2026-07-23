from __future__ import annotations

import hashlib
import json
import math
import posixpath
import re
import xml.etree.ElementTree as ElementTree
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


SOURCE = Path("data/format alur penginputan Accounting.xlsx")
OUTPUT = Path(".tmp/accounting-demo-data.json")
SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
PACKAGE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        compact = re.sub(r"\s+", " ", value).strip()
        return compact or None
    return value


def text(value: Any) -> str | None:
    value = clean(value)
    return str(value) if value is not None else None


def number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    normalized = re.sub(r"[^\d.-]", "", str(value))
    if not normalized:
        return None
    try:
        parsed = float(normalized)
        return parsed if math.isfinite(parsed) else None
    except ValueError:
        return None


def classify(description: str | None) -> str:
    value = (description or "").lower()
    part = bool(re.search(r"spare\s*part|suku cadang|\bpart\b|weld|penjualan", value))
    service = bool(re.search(r"jasa|service|servis|rental|maint|repair|perbaikan|rekondisi|labour|lembur|contract", value))
    if part and service:
        return "mixed"
    if part:
        return "sparepart"
    if service:
        return "service"
    return "other"


def merged_cell_anchors(
    source: Path,
    sheet_name: str,
    *,
    min_row: int,
    max_row: int,
    max_col: int,
) -> dict[tuple[int, int], tuple[int, int]]:
    """Find relevant merged cells without loading the million-row sheet in memory."""
    with zipfile.ZipFile(source) as archive:
        workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationship_id = None
        for sheet in workbook_xml.findall(
            f".//{{{SPREADSHEET_NAMESPACE}}}sheet"
        ):
            if sheet.attrib.get("name") == sheet_name:
                relationship_id = sheet.attrib.get(
                    f"{{{DOCUMENT_RELATIONSHIP_NAMESPACE}}}id"
                )
                break
        if not relationship_id:
            raise KeyError(f"Worksheet {sheet_name!r} was not found")

        relationships_xml = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        target = None
        for relationship in relationships_xml.findall(
            f"{{{PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship"
        ):
            if relationship.attrib.get("Id") == relationship_id:
                target = relationship.attrib.get("Target")
                break
        if not target:
            raise KeyError(f"Worksheet relationship {relationship_id!r} was not found")

        worksheet_path = (
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", target))
        )
        anchors: dict[tuple[int, int], tuple[int, int]] = {}
        with archive.open(worksheet_path) as worksheet_xml:
            for _, element in ElementTree.iterparse(worksheet_xml, events=("end",)):
                if element.tag != f"{{{SPREADSHEET_NAMESPACE}}}mergeCell":
                    element.clear()
                    continue
                reference = element.attrib.get("ref")
                if not reference:
                    element.clear()
                    continue
                min_col_index, min_row_index, max_col_index, max_row_index = (
                    range_boundaries(reference)
                )
                if (
                    max_row_index < min_row
                    or min_row_index > max_row
                    or min_col_index > max_col
                ):
                    element.clear()
                    continue
                anchor = (min_row_index, min_col_index)
                for row_index in range(
                    max(min_row_index, min_row),
                    min(max_row_index, max_row) + 1,
                ):
                    for column_index in range(
                        min_col_index,
                        min(max_col_index, max_col) + 1,
                    ):
                        anchors[(row_index, column_index)] = anchor
                element.clear()
        return anchors


def expand_merged_values(
    source_row: int,
    values: tuple[Any, ...],
    merged_cells: dict[tuple[int, int], tuple[int, int]],
    anchor_values: dict[tuple[int, int], Any],
) -> tuple[Any, ...]:
    expanded = list(values)
    for column_index, value in enumerate(expanded, start=1):
        anchor = merged_cells.get((source_row, column_index))
        if not anchor:
            continue
        if anchor == (source_row, column_index):
            anchor_values[anchor] = value
        elif value is None and anchor in anchor_values:
            expanded[column_index - 1] = anchor_values[anchor]
    return tuple(expanded)


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
rows: list[dict[str, Any]] = []
invoices: list[dict[str, Any]] = []


invoice_sheet = workbook["rek. inv"]
seen_invoice_numbers: set[str] = set()
for source_row, values in enumerate(
    invoice_sheet.iter_rows(min_row=8, max_row=7300, max_col=16, values_only=True),
    start=8,
):
    customer = text(values[1])
    invoice_number = text(values[5])
    subtotal = number(values[10])
    description = text(values[9])
    if not customer or not invoice_number or subtotal is None or subtotal <= 0:
        continue
    tax = number(values[11]) or 0
    total = number(values[12])
    if total is None or total <= 0:
        total = subtotal + tax
    category = classify(description)
    data = {
        "customer": customer,
        "deliveryNoteNumber": text(values[2]),
        "deliveryNoteDate": clean(values[3]),
        "receiptNumber": text(values[4]),
        "invoiceNumber": invoice_number,
        "invoiceDate": clean(values[6]),
        "purchaseOrderNumber": text(values[7]),
        "purchaseOrderDate": clean(values[8]),
        "description": description or "Invoice Accounting",
        "subtotal": subtotal,
        "taxAmount": tax,
        "total": total,
        "taxInvoiceNumber": text(values[13]),
        "accountDestination": text(values[14]),
        "colorCode": text(values[15]),
        "category": category,
    }
    rows.append({"sheetKey": "invoice_register", "sourceRow": source_row, "data": data})
    if invoice_number in seen_invoice_numbers:
        continue
    seen_invoice_numbers.add(invoice_number)
    invoices.append(data | {"sourceRow": source_row})


delivery_sheet = workbook["rekap SJ ( dari Logistik)"]
delivery_merged_cells = merged_cell_anchors(
    SOURCE,
    delivery_sheet.title,
    min_row=7,
    max_row=17066,
    max_col=11,
)
delivery_anchor_values: dict[tuple[int, int], Any] = {}
for source_row, values in enumerate(
    delivery_sheet.iter_rows(min_row=7, max_row=17066, max_col=11, values_only=True),
    start=7,
):
    values = expand_merged_values(
        source_row,
        values,
        delivery_merged_cells,
        delivery_anchor_values,
    )
    company = text(values[0])
    delivery_number = text(values[1])
    if not any(clean(value) is not None for value in values):
        continue
    data = {
        "company": company,
        "deliveryNoteNumber": delivery_number,
        "deliveryNoteDate": clean(values[2]),
        "invoiceNumber": text(values[3]),
        "invoiceDate": clean(values[4]),
        "purchaseOrderNumber": text(values[5]),
        "purchaseOrderDate": clean(values[6]),
        "description": text(values[7]),
        "subtotal": number(values[8]),
        "taxAmount": number(values[9]),
        "total": number(values[10]),
    }
    rows.append({"sheetKey": "delivery_notes", "sourceRow": source_row, "data": data})


receivable_sheet = workbook["rek. penapatan inv. jasa & part"]
months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
for source_row, values in enumerate(
    receivable_sheet.iter_rows(min_row=10, max_row=164, max_col=18, values_only=True),
    start=10,
):
    customer = text(values[1])
    if not customer:
        continue
    data = {
        "number": number(values[0]),
        "customer": customer,
        "totalReceivable": number(values[2]) or 0,
        "paid": number(values[3]) or 0,
        "balance": number(values[4]) or 0,
        "months": {
            month: number(values[index + 5]) or 0
            for index, month in enumerate(months)
        },
        "notes": text(values[17]),
    }
    rows.append({"sheetKey": "invoice_receivables", "sourceRow": source_row, "data": data})


part_sheet = workbook["pend. part ( sesuai invoice)"]
for source_row, values in enumerate(
    part_sheet.iter_rows(min_row=8, max_row=182, max_col=13, values_only=True),
    start=8,
):
    customer = text(values[1])
    invoice_number = text(values[5])
    if not customer or not invoice_number:
        continue
    data = {
        "customer": customer,
        "deliveryNoteNumber": text(values[2]),
        "deliveryNoteDate": clean(values[3]),
        "receiptNumber": text(values[4]),
        "invoiceNumber": invoice_number,
        "invoiceDate": clean(values[6]),
        "purchaseOrderNumber": text(values[7]),
        "purchaseOrderDate": clean(values[8]),
        "subtotal": number(values[9]) or 0,
        "taxAmount": number(values[10]) or 0,
        "total": number(values[11]) or 0,
        "taxInvoiceNumber": text(values[12]),
    }
    rows.append({"sheetKey": "spare_part_income", "sourceRow": source_row, "data": data})


service_sheet = workbook["pend. jasa ( sesuai invoice)"]
for source_row, values in enumerate(
    service_sheet.iter_rows(min_row=9, max_row=68, max_col=12, values_only=True),
    start=9,
):
    customer = text(values[1])
    invoice_number = text(values[3])
    if not customer or not invoice_number:
        continue
    data = {
        "customer": customer,
        "receiptNumber": text(values[2]),
        "invoiceNumber": invoice_number,
        "invoiceDate": clean(values[4]),
        "purchaseOrderNumber": text(values[5]),
        "purchaseOrderDate": clean(values[6]),
        "subtotal": number(values[7]) or 0,
        "taxAmount": number(values[8]) or 0,
        "total": number(values[9]) or 0,
        "taxInvoiceNumber": text(values[10]),
        "notes": text(values[11]),
    }
    rows.append({"sheetKey": "service_income", "sourceRow": source_row, "data": data})


summary_sheet = workbook["pend. jasa & part"]
for source_row, values in enumerate(
    summary_sheet.iter_rows(min_row=8, max_row=79, max_col=7, values_only=True),
    start=8,
):
    customer = text(values[1])
    if not customer:
        continue
    data = {
        "customer": customer,
        "partIncome": number(values[2]) or 0,
        "serviceIncome": number(values[3]) or 0,
        "combinedIncome": number(values[4]) or 0,
        "taxAmount": number(values[5]) or 0,
        "total": number(values[6]) or 0,
    }
    if not any(value for key, value in data.items() if key != "customer"):
        continue
    rows.append({"sheetKey": "service_part_summary", "sourceRow": source_row, "data": data})


contract_sheet = workbook["kontrak"]
last_customer: str | None = None
for source_row, values in enumerate(
    contract_sheet.iter_rows(min_row=7, max_row=167, max_col=14, values_only=True),
    start=7,
):
    customer = text(values[1])
    if customer:
        last_customer = customer
    contract_number = text(values[3])
    period = text(values[5])
    if not last_customer or not any(text(value) for value in values[2:12]):
        continue
    data = {
        "number": text(values[0]),
        "customer": last_customer,
        "vendor": text(values[2]),
        "contractNumber": contract_number,
        "signatory": text(values[4]),
        "period": period,
        "contractValue": number(values[6]),
        "additionalValue": number(values[7]),
        "mechanicCount": number(values[8]),
        "mechanicUnit": text(values[9]),
        "workingHours": text(values[10]),
        "remarks": text(values[11]),
        "raw": [clean(value) for value in values],
    }
    rows.append({"sheetKey": "contracts", "sourceRow": source_row, "data": data})


source_hash = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
counts: dict[str, int] = {}
for row in rows:
    counts[row["sheetKey"]] = counts.get(row["sheetKey"], 0) + 1

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(
    json.dumps(
        {
            "sourceFileName": SOURCE.name,
            "sourceSha256": source_hash,
            "counts": counts,
            "invoices": invoices,
            "rows": rows,
        },
        ensure_ascii=False,
        indent=2,
    ),
    encoding="utf-8",
)
print(json.dumps({"output": str(OUTPUT), "counts": counts, "invoices": len(invoices)}))
